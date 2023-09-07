from pathlib import Path
from openpyxl import load_workbook, Workbook
import openpyxl as xl
import os
import pandas as pd
import glob
os.system("cls")

# SOURCE_DIR=r"C:\Users\ke5ol\Documents\norman property list\prop_list_downloads"
# excel_files=list(Path(SOURCE_DIR).glob("*.xlsx"))


# specifying the path to csv files
path = r"C:\Users\ke5ol\Documents\norman property list\prop_list_downloads"
# csv files in the path
file_list = glob.glob(path + "/*.xlsx")
 
# list of excel files we want to merge.
# pd.read_excel(file_path) reads the excel
# data into pandas dataframe.
excl_list = []
 
for file in file_list:
    excl_list.append(pd.read_excel(file))
 
# create a new dataframe to store the
# merged excel file.
excl_merged = pd.DataFrame()
 
for excl_file in excl_list:
     
    # appends the data into the excl_merged
    # dataframe.
    excl_merged = excl_merged.append(
      excl_file, ignore_index=True)
 
# exports the dataframe into excel file with
# specified name.
excl_merged.to_excel('merged_prop_list.xlsx', index=False)












# values_excel_files={}
# filename="combined_properties.xlsx"
# with pd.ExcelWriter(filename, mode='w',engine='openpyxl') as writer:
	# df = pd.DataFrame(columns=['Account #','Owner','Address','Market Value','Parcel ID','Website'])
	# df.to_excel(writer,index=False)


# for excel_file in excel_files:
	# wb=load_workbook(filename=excel_file)
	# ws=wb.active
	# mr = ws.max_row
	# # print(excel_file)
	# # print(mr)
	# # print(mc)
	# # print("F"+str(mr))
	# maxCell="F"+str(mr)
	# rng=wb["Real Property Record Data"]["B2":maxCell]
	# rng_values=[]
	# for cells in rng:
		# for cell in cells:
			# rng_values.append(cell.value)
			# values_excel_files[excel_file]=rng_values
	# df1=pd.DataFrame.from_dict(values_excel_files, orient='index')
	# with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
		# df1.to_excel(writer,sheet_name="Sheet1",index=False)
			



# with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
	# df1.to_excel(writer,sheet_name="Sheet1",index=False)
# wb1 = xl.load_workbook(filename)
# wb1.append(values_excel_files.value)
# wb1.save(str(filename))

# print(values_excel_files)
# wb=load_workbook(filename="combined_properties.xlsx")
# for ws in wb.worksheets:
	# clm="A"
	# first_row=2
	# last_row=len(ws[clm])
	# rng=ws[f"{clm}{first_row}:{clm}{last_row}"]
	# for cells in rng:
		# for cell in cells: